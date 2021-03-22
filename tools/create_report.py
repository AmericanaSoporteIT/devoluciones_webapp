

from flask import send_file, Response
from flask import Blueprint
from openpyxl import workbook
import pandas as pd 
import os
from datetime import *

from Application.connect_server import *
from Application.breadcrumb import breadcrumb

 
    
bp = Blueprint('reports', __name__, url_prefix='/reports')

@bp.route('/', methods=('GET', 'POST'))
@breadcrumb('Reportes')
def report():
    conn = Connection()
    
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    
    today = BASE_DIR+os.sep+"reports"+ os.sep + datetime.now().strftime('%Y%m%d_%H%M%S')
    if not os.path.exists(today):
        os.mkdir(today)

    P_data = pd.read_sql("SELECT * FROM USUARIOS", conn)
    P_data.to_excel(today+os.sep+"excel"+".xlsx")
    
    file = today+os.sep+"excel.xlsx"
        # target = open(today+os.sep+"excel.xlsx", 'rb') 
        # file = target.read()
        # target.close()
    #target = pd.read_excel(file)
    from openpyxl import Workbook
    workbook_errors = Workbook()
    workbook_errors = openpyxl.load_workbook(file)
    
    return send_file(target,mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True , attachment_filename='file.xlsx')



# from flask import send_file
# from xlsxwriter import Workbook

# @app.route('/download', methods=['GET'])
# def export_db():
#     values = execute("SELECT * from table",[])
#     wb = Workbook('path/to/workbook.xlsx')
#     wb.add_worksheet('All Data')

#     for item in values.fetchall():
#         wb.write(item)
#     wb.close()

#     return send_file('path/to/workbook.xlsx')


# result = "SELECT * FROM OE_TAT where convert(date,Time_IST)=?"
# df = pd.read_sql_query(result, connection, params=(dateval,))   
# @app.route('/show/static-pdf/')
# def show_static_pdf():
#     with open('/path/of/file.pdf', 'rb') as static_file:
#         return send_file(static_file, attachment_filename='file.pdf')


# from flask import send_from_directory, current_app as app

# @app.route('/show/PDFs/')
# def send_pdf():
#     return send_from_directory(app.config['UPLOAD_FOLDER'], 'file.pdf')